import openpyxl

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import os
import csv

pathway = r'C:\Users\user\Desktop\renamed'
list_of_files=os.listdir(pathway)


# create dictionaries to for each of the variables
names = ['Customer name']
amount = ['DIAGNOSIS AMOUNT']


def customer_name(row_no):
    pattern = r'(\d+)' # introducing the rgex function
    l= re.split(pattern, row_no)# split the pattern of the regex functions

    cell = 'B' + str(l[1]) #this contactenates the regex functions to the cell value
    print(cell)
    record=(wb[cell].value)
    print(f'customer name:{record}')
    if record == None:
        names.append('NIL')
    else:
        names.append(record)
    return names


'''this function saves the phone numbers'''

def amounts(data):
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    check_PRO=int(l[1])
    cell = 'F' + str(check_PRO)
    print(cell)
    record = wb[cell].value
    print(f' diagnosis amount:{record}')
    if record == None:
        amount.append('NIL')
    else:
        amount.append(record)
    return amount

def open_new_workbook():
    workbook= Workbook()
    w= workbook.create_sheet('worksheet 1')
    ws=workbook.active

#paste the scraped data in the new excel sheet'''
    for a,b in zip(names, amount):
        list=a,b
        #print(list)
        ws.append(list)
    workbook.save(r'C:\Users\user\Desktop\mad\new_month.xlsx')

open_new_workbook()







'''iterate over the cells and pass the cell values to the appropriate functions above'''
for n in list_of_files:
    if n != 'desktop.ini':
        print(n)
        without_being_told = pathway + f'\{n}'
        print(without_being_told)
        path = without_being_told



        wb_obj = load_workbook(path)
        for sheet in wb_obj.worksheets:
            print(sheet.title)
            wb = wb_obj[str(sheet.title)]
            print(wb)
            for row in range(1, 2000):
                for col in range(1, 500):

                    char = get_column_letter(col)
                    search = char + str(row)
                    search2 = wb[char + str(row)].value
                    if search2 == "DIAGNOSIS":
                        print(search)
                        print(search2)
                        customer_name(search)
                        amounts(search)








'''print all the process to be sure of the process being undergone'''

print(f'number of names{len(names)}') #print the lenth of the list of names
print(f'amounts: {amount}') #print all the phone numbers
collection=[n for n in amount if n == 'NIL']
print(len(collection))
print(collection)
print(f'CUSTOMER NAME;{names}')



'''create a function that opens the workbook and makes it active'''

# now open a new workbook
# save/paste the data in the new workbook


def new_function():
    print(open_new_workbook())


new_function()
