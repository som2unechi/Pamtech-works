import os
import re
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def extract_vehicle_number(cell_value):
    pattern = r'(\d+)'
    result = re.findall(pattern, cell_value)
    return result[0] if result else 'NIL'


def extract_customer_name(cell_value):
    pattern = r'(\d+)'
    result = re.findall(pattern, cell_value)
    return result[0] if result else 'NIL'


def extract_date(cell_value):
    pattern = r'(\d+)'
    result = re.findall(pattern, cell_value)
    return result[0] if result else 'NIL'


def scrape_excel_files(directory):
    extracted_data = []

    for root, _, files in os.walk(directory):
        for file_name in files:
            if file_name != 'desktop.ini':
                file_path = os.path.join(root, file_name)
                workbook = load_workbook(file_path)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    rows = list(sheet.iter_rows(values_only=True))

                    for row_index, row in enumerate(rows, start=1):
                        for cell_index, cell_value in enumerate(row, start=1):
                            if cell_value == "APPROVED":
                                contact_person_cell = None
                                customer_name_cell = None
                                for prev_row_index in range(row_index - 1, 0, -1):
                                    prev_row = rows[prev_row_index - 1]
                                    for prev_cell_value in prev_row:
                                        if prev_cell_value == "Contact Person:":
                                            contact_person_cell = sheet.cell(row=prev_row_index, column=prev_row.index(prev_cell_value) + 1)
                                            break
                                        elif prev_cell_value == "Customer Name:":
                                            customer_name_cell = sheet.cell(row=prev_row_index, column=prev_row.index(prev_cell_value) + 1)
                                            break
                                    if contact_person_cell and customer_name_cell:
                                        break

                                if contact_person_cell and customer_name_cell:
                                    contact_person = sheet.cell(row=contact_person_cell.row, column=contact_person_cell.column + 1).value
                                    customer_name = extract_customer_name(str(customer_name_cell.value))
                                    phone_number = extract_phone_number(str(row[0]))
                                    vehicle_number = extract_vehicle_number(str(row[0]))
                                    date = extract_date(str(row[0]))

                                    extracted_data.append([customer_name, phone_number, vehicle_number, date, contact_person])

    return extracted_data


def save_data_to_excel(data, save_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scraped Data"

    for row in data:
        sheet.append(row)

    workbook.save(save_path)


def save_data_to_csv(data, save_path):
    with open(save_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(data)


def main():
    directory_path = r'C:\Users\user\Desktop\virgin'
    output_excel_path = r'C:\Users\user\Desktop\New folder\MARCH SCRAPER.xlsx'
    output_csv_path = 'NEWAUTOLAND2023.csv'

    scraped_data = scrape_excel_files(directory_path)

    save_data_to_excel(scraped_data, output_excel_path)
    save_data_to_csv(scraped_data, output_csv_path)


if __name__ == "__main__":
    main()
