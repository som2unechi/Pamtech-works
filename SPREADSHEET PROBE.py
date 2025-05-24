''' this is a script that scrapes the excel workbook of a company's data sheet to extract  client's data
such as the names, phone numbers, vehicle numbers and make a database off it

 Author: Somtoochukwu Chukwueze'''
import openpyxl

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import os
import csv
from openpyxl.worksheet import worksheet

''' 1. create pathway directories to save the Excel files you wish to scrape
    2. write the pathway directory for python to read through the os library'''

pathway = r'C:\Users\user\Desktop\renamed'
list_of_files=os.listdir(pathway)


# create dictionaries to for each of the variables
names = ['Customer name']
phone = ['VEHICLE TYPE']
reg_no = ['Vehicle number']
dates = ['date']
emails=['email']
complaint=['complaints']



#details= wb.iter_columns(max_column= 1, max_row=1)


# create functions for vehicle number
def vehicle_no(data):
    # use regex to get the cell value
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    '''the cocatenate_your_ass library is just a fancy name that adds/concatenates the regex value to the numerical value of the cell. e.g A+1= A1'''

    concatenate_your_ass = int(l[1])+ 7

    cell = 'E' + str(concatenate_your_ass)
    print(cell)
    record = (wb[cell].value)
    print(f'reg_no:{record}')
    '''The if and else commaand below here is create a condition where the selected cell has no value so it's to return 'none'''
    if record == None:
        reg_no.append('NIL')
    else:
        reg_no.append(record)  # this appends the record of the cell value to the dictionary i created earlier above
    return reg_no

'''creat functions for customer_name'''
def customer_name(row_no):
    pattern = r'(\d+)' # introducing the rgex function
    l= re.split(pattern, row_no)# split the pattern of the regex functions

    cell = 'B' + str(l[1])  #this contactenates the regex functions to the cell value
    print(cell)
    record=(wb[cell].value)
    print(f'customer name:{record}')
    if record == None:
        names.append('NIL')
    else:
        names.append(record)
    return names

def date(data):

    pattern = r'(\d+)'
    l= re.split(pattern, data)

    cell = 'B' + str(int(l[1])+1)
    print(cell)
    record=(wb[cell].value)
    print(f'customer name:{record}')
    if record == None:
        dates.append('NIL')
    else:
        dates.append(record)
    return dates


'''this function saves the phone numbers'''

def phone_number(data):
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    check_PRO=int(l[1])+ 5
    cell = 'B' + str(check_PRO)
    print(cell)
    record = wb[cell].value
    print(f'phone number:{record}')
    if record == None:
        phone.append('NIL')
    else:
        phone.append(record)
    return phone


def complaints(data):
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    cell = 'A' + str(int(l[1]) + 9)
    print(cell)
    record = (wb[cell].value)
    print(f'customer name:{record}')

    if record == None:

        complaint.append('NIL')
    else:
        complaint.append(record)
    return complaint


def email(data, wb):
    pattern = r'(\d+)'
    l = re.split(pattern, data)
    check_PRO = int(l[1]) + 8
    cell = 'B' + str(check_PRO)
    print(cell)
    record = wb[cell].value
    print(f'this is cell cordinate{record}')

    #omo=wb.cell(l[0],l[1])
    #cell_nemesis=wb.cell(row=l[0], column=l[1])
    print(f'email:{record}')
    if record is None:
        emails.append('NIL')
    else:
        emails.append(record)



'''iterate over the cells and pass the cell values to the appropriate functions above'''
for n in list_of_files:
    if n != 'desktop.ini':
        print(n)
        without_being_told = pathway + f'\{n}'
        print(without_being_told)
        path = without_being_told


        wb_obj = load_workbook(path)
        for sheet in wb_obj.worksheets:
            print(sheet.title)
            wb = wb_obj[str(sheet.title)]
            print(wb)
            for row in range(1, 2000):
                for col in range(1, 500):

                    char = get_column_letter(col)
                    search = char + str(row)
                    search2 = wb[char + str(row)].value
                    if search2 == "Customer Name:":
                        print(search)
                        print(search2)
                        customer_name(search)
                        phone_number(search)
                        vehicle_no(search)
                        date(search)
                        email(search,wb)
                        complaints(search)

'''print all the process to be sure of the process being undergone'''
print(f'Reg no: {reg_no}')# print all the reg no
print(f'number of reg_no {len(reg_no)}') #print the length of the list of the reg no
print(f'number of names{len(names)}') #print the lenth of the list of names
print(f'Phone no: {phone}') #print all the phone numbers
print(f'number of phone numbers{len(phone)}') #print the length of the phone numbers list
print(f'dates{dates}') #print the list of dates
collection=[n for n in phone if n == 'NIL']
print(len(collection))
print(collection)
print(f'CUSTOMER NAME;{names}')
print(f'emails: {emails}')
print(f'number of emails dictionary:{len(emails)}')
print(f'the customer complaints{complaint}')
print(f'the customer complaints {len(complaint)} ')

'''create a function that opens the workbook and makes it active'''
def open_new_workbook():
    workbook= Workbook()
    w= workbook.create_sheet('worksheet 1')
    ws=workbook.active

#paste the scraped data in the new excel sheet'''
    for a,b,n,x, y,z in zip(names, phone, reg_no, dates, emails, complaint):
        list=a,b,n,x,y,z
        #print(list)
        ws.append(list)
    workbook.save(r'C:\Users\user\Desktop\mad\august.xlsx')

open_new_workbook()

# now open a new workbook
# save/paste the data in the new workbook


def new_function():
    print(open_new_workbook())


new_function()

with open('../NEWAUTOLAND2023.csv', 'w', newline='') as new_file:
    # Create a writer object
    writer = csv.writer(new_file)
    # Write the header row
    writer.writerow(phone)