#CODE 1; PROBLEM: OVERWRITING THE PRE EXISTING WORKBOOK



#import openpyxl
# import os
#
# pathway = r'C:\Users\user\Desktop\virgin'
# list_of_files= os.listdir(pathway)
# for n in list_of_files:
#     if n != 'desktop.ini':
#         print(n)
#         without_being_told = pathway + f'\{n}'
#         print(without_being_told)
#         path = without_being_told
#
#         # Load the workbook
#         workbook = openpyxl.load_workbook(path)
#
#         # Loop through all sheets in the workbook
#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]
#             print(f'Sheet: {sheet_name}')
#
#             # Loop through all cells in the sheet and print their values
#             for row in sheet.iter_rows():
#                 for cell in row:
#                     print(cell.value)
#
#
#         # Save any changes made to the workbook
#
#         workbook.save('C:/Users/user/Desktop/virgin/new_workbook.xlsx')



# CODE 2; PROBLEM :
# # Load the source workbook
# source_workbook = openpyxl.load_workbook('C:/Users/user/Desktop/virgin/new_workbook.xlsx')
#
# # Select the source worksheet
# for sheet_name in source_workbook.sheetnames:
#     source_worksheet = source_workbook[sheet_name]
#     print(f'Sheet: {sheet_name}')
#
#     # Load the target workbook
#     target_workbook = openpyxl.load_workbook('C:/Users/user/Desktop/virgin/FEBUARY 2023.xlsx')
#
#     # Select the target worksheet
#     target_worksheet = target_workbook['Sheet1']
#
#     # Loop through the rows in the source worksheet and write them to the target worksheet
#     for row in source_worksheet.iter_rows():
#         # Create a list of values for each cell in the row
#         row_values = [cell.value for cell in row]
#
#         # Append the row to the target worksheet
#         target_worksheet.append(row_values)
#
#     # Save the updated target workbook
#     target_workbook.save('C:/Users/user/Desktop/virgin/FEBUARY 2023.xlsx')


#CODE 3: PROBLEM; NOT RETAINING CELL FORMATTING.
# import openpyxl
# import os
#
# pathway = r'C:\Users\user\Desktop\virgin'
# new_workbook = openpyxl.Workbook()  # create a new workbook
# for n in os.listdir(pathway):
#     if n != 'desktop.ini':
#         print(n)
#         without_being_told = pathway + f'\{n}'
#         print(without_being_told)
#         path = without_being_told
#
#         # Load the workbook
#         workbook = openpyxl.load_workbook(path)
#
#         # Loop through all sheets in the workbook
#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]
#             print(f'Sheet: {sheet_name}')
#             new_sheet = new_workbook.create_sheet(title=sheet_name)  # create a new sheet in the new workbook
#             for row in sheet.iter_rows():
#                 for cell in row:
#                     new_sheet[cell.coordinate].value = cell.value  # copy cell values to the new sheet
#
#         # Save any changes made to the new workbook
#         new_workbook.save('C:/Users/user/Desktop/virgin/new_workbook.xlsx')

#
# import openpyxl
# import os
#
# pathway = r'C:\Users\user\Desktop\virgin'
# new_workbook = openpyxl.Workbook()
#
# # Dictionary to map styles to unique identifiers
# style_map = {}
#
# for n in os.listdir(pathway):
#     if n != 'desktop.ini':
#         print(n)
#         without_being_told = pathway + f'\{n}'
#         print(without_being_told)
#         path = without_being_told
#
#         # Load the workbook
#         workbook = openpyxl.load_workbook(path)
#
#         # Loop through all sheets in the workbook
#         for sheet_name in workbook.sheetnames:
#             sheet = workbook[sheet_name]
#             print(f'Sheet: {sheet_name}')
#
#             # Create new sheet in new workbook
#             new_sheet = new_workbook.create_sheet(title=sheet_name)
#
#             # Loop through all cells in the sheet and copy values and styles
#             for row in sheet.iter_rows(values_only=True):
#                 new_row = []
#                 for cell in row:
#                     if cell.style not in style_map:
#                         # Generate unique identifier for new style
#                         style_id = str(len(style_map))
#                         style_map[cell.style] = style_id
#                         new_workbook.add_named_style(cell.style, style_id)
#                     else:
#                         style_id = style_map[cell.style]
#
#                     source_cell = sheet.cell(row=cell.row, column=cell.column)
#                     new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
#                     new_cell.style = style_id
#
#         # Save any changes made to the workbook
#         new_workbook.save('C:/Users/user/Desktop/virgin/new_workbook.xlsx')


import openpyxl
import os
from copy import copy

pathway = r'C:\Users\user\Desktop\UNDONE'
output_file ='C:/Users/user/Desktop/AUTOLAND 2024COMPILED_WORKBOOK(JULY UPDATE).xlsx'
# 'C:/Users/user/Desktop/virgin/compiled_workbook.xlsx'
# Create a new workbook to compile all the data
output_workbook = openpyxl.Workbook()

for n in os.listdir(pathway):
    if n != 'desktop.ini':
        print(n)
        without_being_told = pathway + f'\{n}'
        print(without_being_told)
        path = without_being_told

        # Load the workbook
        workbook = openpyxl.load_workbook(path)

        # Loop through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f'Sheet: {sheet_name}')

            # Create a new sheet in the output workbook with the same name as the original sheet
            new_sheet = output_workbook.create_sheet(title=sheet_name)

            # Loop through all cells in the sheet and copy them to the new sheet
            for row in sheet.iter_rows():
                for cell in row:
                    source_cell = sheet.cell(row=cell.row, column=cell.column)
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if source_cell.has_style:
                        new_cell.font = copy(source_cell.font)
                        new_cell.border = copy(source_cell.border)
                        new_cell.fill = copy(source_cell.fill)
                        new_cell.number_format = copy(source_cell.number_format)
                        new_cell.protection = copy(source_cell.protection)
                        new_cell.alignment = copy(source_cell.alignment)

        # Save any changes made to the output workbook
        output_workbook.save(output_file)

