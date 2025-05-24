''' this is a script that scrapes the excel workbook of a company's data sheet to extract  client's data
: names and emails and make a database off it

 Author: Somtoochukwu Chukwueze '''

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import os
import csv

''' 1. create pathway directories to save the Excel files you wish to scrape
    2. write the pathway directory for python to read through the os library'''

pathway = r'C:\Users\user\Desktop\AUTOLAND 2023'
list_of_files = os.listdir(pathway)

# create dictionaries to for each of the variables
emails = ['email']
phone = ['phone numbers']


def email(row_no):
    pattern = r'(\d+)'  # introducing the rgex function
    l = re.split(pattern, row_no)  # split the pattern of the regex functions
    check_PRO = int(l[1])
    cell = 'B' + str(check_PRO)  # this contactenates the regex functions to the cell value
    print(cell)
    record = (wb[cell].value)
    print(f'email:{record}')
    emails.append(record)

    return emails


def phone_number(row_no):
    pattern = r'(\d+)'  # introducing the rgex function
    l = re.split(pattern, row_no)  # split the pattern of the regex functions

    cell = 'B' + str(l[1])  # this contactenates the regex functions to the cell value
    print(cell)
    record = (wb[cell].value)
    print(f'phone numbers:{record}')
    if record == None:
        phone.append('NIL')
    else:
        phone.append(record)
    return phone


for n in list_of_files:
    if n != 'desktop.ini':
        print(n)
        without_being_told = pathway + f'\{n}'
        print(without_being_told)
        path = without_being_told

        wb_obj = load_workbook(path)
        for sheet in wb_obj.worksheets:
            print(sheet.title)
            wb = wb_obj[str(sheet.title)]
            print(wb)
            for row in range(1, 1000):
                for col in range(1, 20):

                    char = get_column_letter(col)
                    search = char + str(row)
                    search2 = wb[char + str(row)].value

                    if isinstance(search2, str) and search2.lower().startswith('emai'):
                        customer_name_cell = find_customer_name(sheet, total_incl_vat_cell.row)

                        email(search)


                    elif search2 == "Contact Person:":

                        phone_number(search)

    else:
        print('none')

'''print all the process to be sure of the process being undergone'''
print(f'emails: {emails}')  # print all the reg no
print(f'number of emails {len(emails)}')  # print the length of the list of the reg no
print(f'number of phone numbers {len(phone)}')  # print the lenth of the list of phone numbers
print(f'Phone no: {phone}')  # print all the phone numbers

'''create a function that opens the workbook and makes it active'''


def open_new_workbook():
    workbook = Workbook()
    w = workbook.create_sheet('worksheet 1')
    ws = workbook.active

    # paste the scraped data in the new excel sheet'''
    for x, y in zip(emails, phone):
        list = x, y
        # print(list)

        ws.append(list)

    workbook.save('C:/Users/user/Desktop/openpyxl files folder/PRAISES ASSIGNMENT.xlsx')


open_new_workbook()


def new_function():
    print(open_new_workbook())


new_function()

with open('praisesemail.csv', 'w', newline='') as new_file:
    # Create a writer object
    writer = csv.writer(new_file)
    # Write the header row
    writer.writerow(phone)
    writer.writerow(emails)

