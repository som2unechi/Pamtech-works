from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import os
import csv

pathway = r'C:\Users\user\Desktop\renamed'
list_of_files=os.listdir(pathway)


# continue from  here tommorrow 26/08/22 ; try to seperate workshseet from data; consider using regex again.
names = ['Customer name']
phone = ['Phone Numbers']
reg_no = ['Vehicle number']
TOTAL = ['TOTAL INCLUDING VAT']

#details= wb.iter_columns(max_column= 1, max_row=1)

"""cell_obj = sheet_obj.cell(row = 1, column = 1)
cell_obj2 = sheet_obj.cell(row = 1, column = 2)

print(cell_obj.value,cell_obj2.value)
print(sheet_obj)

max_col= sheet_obj.max_column
for i in range (1, max_col + 1):
    cell_obj3=sheet_obj.cell( row=28, column = i)
    print (cell_obj3.value, end=" """


"""def total(data):
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    concatenate_your_ass = int(l[1])
    cell = 'G' + str(concatenate_your_ass)
    print(cell)

    record = (wb[cell].internal_value)
    print(f'TOTAL AMOUNT:{record}')
    if record == None:
        TOTAL.append('NIL')
    else:
        TOTAL.append(record)
    return TOTAL"""



def vehicle_no(data):

    pattern = r'(\d+)'
    l = re.split(pattern, data)

    concatenate_your_ass = int(l[1])+8

    cell = 'E' + str(concatenate_your_ass)
    print(cell)
    record = (wb[cell].value)
    print(f'TOTAL:{record}')
    if record == None:
        reg_no.append('NIL')
    else:
        reg_no.append(record)
    return reg_no


def customer_name(row_no):
    pattern = r'(\d+)'
    l= re.split(pattern, row_no)

    cell = 'B' + str(l[1])
    print(cell)
    record=(wb[cell].value)
    print(f'customer name:{record}')
    if record == None:
        names.append('NIL')
    else:
        names.append(record)
    return names

def phone_number(data):
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    check_PRO=int(l[1])+2
    cell = 'B' + str(check_PRO)
    print(cell)
    record = wb[cell].value
    print(f'phone number:{record}')
    if record == None:
        phone.append('NIL')
    else:
        phone.append(record)


    return phone


for n in list_of_files:
    if n != 'desktop.ini':
        print(n)
        without_being_told = pathway + f'\{n}'
        print(without_being_told)
        path = without_being_told

        wb_obj = load_workbook(path, data_only=True)
        for sheet in wb_obj.worksheets:
            print(sheet.title)
            wb = wb_obj[str(sheet.title)]
            print(wb)
            for row in range(1, 1000):
                for col in range(1, 20):
                    # val=wb[col]
                    # for n in val:
                    #     print(n.value)
                    #    print(val)+

                    char = get_column_letter(col)
                    search = char + str(row)
                    search2 = wb[char + str(row)].value
                    if search2 == "Customer Name:":
                        print(search)
                        print(search2)
                        customer_name(search)
                        phone_number(search)
                        vehicle_no(search)

print(f'Reg no: {reg_no}')
print(f'number of reg_no {len(reg_no)}')
print(f'number of names{len(names)}')
print(f'Phone no: {phone}')
print(f'number of phone numbers{len(phone)}')
collection=[n for n in phone if n == 'NIL']


print(f'CUSTOMER NAME;{names}')


def open_new_workbook():
    workbook= Workbook()
    w= workbook.create_sheet('worksheet 1')
    ws=workbook.active

    for n,x, y in zip(names, phone, reg_no):
        list=n,x,y
        #print(list)


        ws.append(list)

        #ws.append()

    #for row in range(1, 300):
    #    ws['A1']='why'
     #   print(ws['A1'].value)
      #  ws.append(names)
    workbook.save('C:/Users/user/Desktop/New folder/MARCH SCRAPER.xlsx')


open_new_workbook()
data = ['John', 'Jane', 'Jim']

# File name
filename = 'example.csv'

# Open the file in write mode
with open(filename, 'w', newline='') as file:
    # Create a writer object
    writer = csv.writer(file)

    # Write the data

    writer.writerow(phone)



# Open the CSV file in read mode


# Open the CSV file in read mode
with open('example.csv', 'r') as file:
    # Create a reader object
    reader = csv.reader(file)
    # Skip the header row
    next(reader)
    # Open a new file to write the cleaned phone numbers
    with open('cleaned_phone_numbers.csv', 'w', newline='') as new_file:
        # Create a writer object
        writer = csv.writer(new_file)
        # Write the header row
        writer.writerow(['Phone Numbers'])
        # Iterate over each row in the CSV file
        for row in reader:
            # Replace all spaces in the phone number with an empty string
            phone_number = row[0].replace(' ', '')
            # Write the cleaned phone number to the new file
            writer.writerow([phone_number])



# now open a new workbook
# save/paste the data in the new workbook


