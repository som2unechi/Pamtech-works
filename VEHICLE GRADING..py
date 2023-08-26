import openpyxl

import openpyxl

# def scrape_spreadsheet(file_path):
#     # Load the workbook
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#
#     # Define dictionaries for each vehicle make
#     benz_dict = {}
#     lexus_dict = {}
#     toyota_dict = {}
#
#     # Iterate through the cells in the sheet
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell and str(cell).startswith("Model"):
#                 print(cell.value)
#                 # Check the next cell for the make
#                 make_cell = row[cell.column + 1]
#                 if make_cell:
#                     make = make_cell.strip().upper()
#                     if make and str(make).startswith('LEXUS'):
#                         # Add data to the BENZ dictionary
#                         benz_dict[row[make_cell.column - 1]] = make_cell.value
#                     elif "LEXUS" in make:
#                         # Add data to the LEXUS dictionary
#                         lexus_dict[row[make_cell.column - 1]] = make_cell.value
#                     elif "TOYOTA" in make:
#                         # Add data to the TOYOTA dictionary
#                         toyota_dict[row[make_cell.column - 1]] = make_cell.value
#
#     # Print the length and values of each dictionary
#     print("Length of BENZ dictionary:", len(benz_dict))
#     print("BENZ dictionary values:")
#     for key, value in benz_dict.items():
#         print(key, ":", value)
#     print("")
#
#     print("Length of LEXUS dictionary:", len(lexus_dict))
#     print("LEXUS dictionary values:")
#     for key, value in lexus_dict.items():
#         print(key, ":", value)
#     print("")
#
#     print("Length of TOYOTA dictionary:", len(toyota_dict))
#     print("TOYOTA dictionary values:")
#     for key, value in toyota_dict.items():
#         print(key, ":", value)
#     print("")
#
# # Specify the file path of the spreadsheet
# file_path = r'C:\Users\user\Desktop\TEST\MAY 2023.xlsx'
#
# # Call the scraping function
# scrape_spreadsheet(file_path)
#


''' this is a script that scrapes the excel workbook of a company's data sheet to extract  client's data
such as the names, phone numbers, vehicle numbers and make a database off it

 Author: Somtoochukwu Chukwueze '''

# import required libraries
# openpyxl; reads Excel files
# load_workbook loads the Excel workbooks
# re is regex library to read the cells e.g A1, B1, etc

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import os
import csv

''' 1. create pathway directories to save the Excel files you wish to scrape
    2. write the pathway directory for python to read through the os library'''

pathway = r'C:\Users\user\Desktop\M'
list_of_files=os.listdir(pathway)


# create dictionaries to for each of the variables
names = ['Customer name']
phone = ['VEHICLE TYPE']
reg_no = ['Vehicle number']
dates = ['date']

# details= wb.iter_columns(max_column= 1, max_row=1)
# create functions for vehicle number
def vehicle_no(data):
    # use regex to get the cell value
    pattern = r'(\d+)'
    l = re.split(pattern, data)
# the cocatenate_your_ass library is just a fancy name that adds/concatenates the regex value to the numerical value of the cell. e.g A+1= A1'''
    concatenate_your_ass = int(l[1])+8

    cell = 'E' + str(concatenate_your_ass)
    print(cell)
    record = (wb[cell].value)
    print(f'reg_no:{record}')

    '''The if and else commaand below here is create a condition where the selected cell has no value so it's to return 'none'''

    if record == None:
        reg_no.append('NIL')
    else:
        reg_no.append(record)  # this appends the record of the cell value to the dictionary i created earlier above
    return reg_no

# '''creat functions for customer_name'''
# def customer_name(row_no):
#     pattern = r'(\d+)' # introducing the rgex function
#     l= re.split(pattern, row_no)# split the pattern of the regex functions
#
#     cell = 'B' + str(l[1])  #this contactenates the regex functions to the cell value
#     print(cell)
#     record=(wb[cell].value)
#     print(f'customer name:{record}')
#     if record == None:
#         names.append('NIL')
#     else:
#         names.append(record)
#     return names
#
# def date(data):
#     pattern = r'(\d+)'
#     l= re.split(pattern, data)
#
#     cell = 'B' + str(int(l[1])+1)
#     print(cell)
#     record=(wb[cell].value)
#     print(f'customer name:{record}')
#     if record == None:
#         dates.append('NIL')
#     else:
#         dates.append(record)
#     return dates


'''this function checks if the cell has the value benz in it'''
benzima=[]
lexus=[]
toyota=[]
hyundai=[]
honda=[]
ford=[]
range_rover=[]
nissan=[]
nil=[]

def benz(data):
    pattern = r'(\d+)'
    l = re.split(pattern, data)

    check_PRO=int(l[1])
    cell = 'B' + str(check_PRO)
    #print(cell)
    record = wb[cell].value
    #print(record)
    if record and str(record).startswith("BENZ" or "M." or "ML" or "GL" or "MERCEDES"):
        #print(record)
        benzima.append(record)
        #print(benzima)
        print(benzima)

    elif str(record).startswith("MERCEDES"):
        benzima.append(record)
        print(benzima)


    elif str(record).startswith("LEXUS"):
        # Add data to the LEXUS dictionary
        lexus.append(record)
        #print(f'appended:{lexus}')
        print(lexus)

    elif str(record).startswith("TOYOTA" or "CAMRY"):
        # Add data to the TOYOTA dictionary
        toyota.append(record)
        #print(f'appended:{toyota}')
        print(toyota)

    elif str(record).startswith("HYUNDAI"):
        # Add data to the TOYOTA dictionary
        hyundai.append(record)
        #print(f'appended:{toyota}')
        print(hyundai)

    elif str(record).startswith("HONDA"):
        # Add data to the TOYOTA dictionary
        honda.append(record)
        #print(f'appended:{toyota}')
        print(honda)

    elif str(record).startswith("FORD"):
        # Add data to the TOYOTA dictionary
        ford.append(record)
        #print(f'appended:{toyota}')
        print(ford)

    elif str(record).startswith("RANGE" or "LAND ROVER"):
        range_rover.append(record)
        print(range_rover)
    elif str(record).startswith("LAND ROVER"):
        range_rover.append(record)
        print(range_rover)

    elif str(record).startswith("NIS" or "INFINI"):
        nissan.append(record)
        print(nissan)
    elif str(record).startswith("INFINI"):
        nissan.append(record)
        print(nissan)

    else:
        nil.append(f'nil{wb[cell].value}')
        print(nil)



'''iterate over the cells and pass the cell values to the appropriate functions above'''
for n in list_of_files:
    if n != 'desktop.ini':
        print(n)
        without_being_told = pathway + f'\{n}'
        #print(without_being_told)
        path = without_being_told

        wb_obj = load_workbook(path)
        for sheet in wb_obj.worksheets:
            #print(sheet.title)
            wb = wb_obj[str(sheet.title)]
            #print(wb)
            for row in range(1, 1000):
                for col in range(1, 20):

                    char = get_column_letter(col)
                    search = char + str(row)
                    search2 = wb[char + str(row)].value
                    if search2 == 'Model Make:':

                        #print(search2)
                        #customer_name(search)
                        benz(search)
        print(f'the list of vehices that came in on {path} is toyota: {len(toyota)},nil;  {len(nil)} : {nil} \n '
              f'nissan: {len(nissan)}ford: {len(ford)}, range rover: {len(range_rover)},'
              f' Hyundai: {len(hyundai)} ,\n'
              f' HOnda : {len(honda)} ,benz: {len(benzima)}, lexus:{len(lexus)}')
                        #vehicle_no(search)
                        #date(search)






def open_new_workbook():
    workbook= Workbook()
    w= workbook.create_sheet('worksheet 1')
    ws=workbook.active

#paste the scraped data in the new excel sheet'''
    for n,x, y,z in zip(names, phone, reg_no, dates):

        list=n,x,y,z
        #print(list)


        ws.append(list)

        #ws.append()

    #for row in range(1, 300):
    #    ws['A1']='why'
     #   print(ws['A1'].value)
    #save the workbook into a path
    workbook.save('C:/Users/user/Desktop/New folder/MARCH SCRAPER2.xlsx')


open_new_workbook()

# now open a new workbook
# save/paste the data in the new workbook


def new_function():
    print(open_new_workbook())


new_function()

with open('NEWAUTOLAND20232.csv', 'w', newline='') as new_file:
    # Create a writer object
    writer = csv.writer(new_file)
    # Write the header row
    writer.writerow(phone)